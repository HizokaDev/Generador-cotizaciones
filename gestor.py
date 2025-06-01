# gestor.py
import os
from openpyxl import Workbook, load_workbook
from datetime import date
from pathlib import Path

import unicodedata
import re


def limpiar_nombre_archivo(nombre: str) -> str:
    # Normalizar (eliminar tildes y otros diacríticos)
    nombre = unicodedata.normalize('NFD', nombre)
    nombre = nombre.encode('ascii', 'ignore').decode('utf-8')
    # Pasar a minúsculas
    nombre = nombre.lower()
    # Reemplazar espacios y otros caracteres no alfanuméricos por guion bajo
    nombre = re.sub(r'[^a-z0-9]+', '_', nombre)
    # Quitar guiones bajos extras al principio y final
    nombre = nombre.strip('_')
    return nombre

def normalizar_texto(texto):
    texto = texto.lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto

archivo = "/storage/emulated/0/DojoHizoka/generador_cotizaciones/cotizaciones.xlsx"

def limpiar():
    os.system("clear")

def crear_archivo_si_no_existe():
    
    carpeta = os.path.dirname(archivo)
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
    ruta = Path(archivo)
    if not ruta.exists():
        wb = Workbook()
        hoja = wb.active
        hoja.append(["Cliente", "Producto", "Cantidad", "Precio Unitario", "Total", "Fecha"])
        wb.save(archivo)

def crear_cotizacion():
    wb = load_workbook(archivo)
    hoja = wb.active

    cliente = input("Nombre del cliente: ")
    hoy = date.today().strftime("%Y-%m-%d")

    while True:
        producto = input("Nombre del producto: ")
        cantidad = int(input("Cantidad: "))
        precio = float(input("Precio unitario: "))
        total = cantidad * precio

        hoja.append([cliente, producto, cantidad, precio, total, hoy])
        wb.save(archivo)

        otro = input("¿Desea agregar otro producto para este cliente? (s/n): ").lower()
        if otro != "s":
            break

    print("✅ Cotización registrada correctamente.")

def buscar_cotizacion():
    cliente = input("Ingrese el nombre del cliente: ")
    cliente_normalizado = normalizar_texto(cliente)

    wb = load_workbook(archivo)
    ws = wb.active

    encontrado = False
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre_archivo = normalizar_texto(fila[0])
        if nombre_archivo == cliente_normalizado:
            print("\nCotización encontrada:")
            print(f"Cliente: {fila[0]}")
            print(f"Producto: {fila[1]}")
            print(f"Cantidad: {fila[2]}")
            print(f"Precio unitario: {fila[3]}")
            print(f"Total: {fila[4]}")
            encontrado = True
            break

    if not encontrado:
        print("No se encontró ninguna cotización para ese cliente.")

def editar_cotizacion():
    cliente = input("Ingrese el nombre del cliente a editar: ")
    cliente_normalizado = normalizar_texto(cliente)

    wb = load_workbook(archivo)
    ws = wb.active
    fila_encontrada = None

    for i, fila in enumerate(ws.iter_rows(min_row=2), start=2):
        nombre_archivo = normalizar_texto(str(fila[0].value))
        if nombre_archivo == cliente_normalizado:
            fila_encontrada = i
            break

    if fila_encontrada:
        nuevo_producto = input("Nuevo producto: ")
        nueva_cantidad = int(input("Nueva cantidad: "))
        nuevo_precio = float(input("Nuevo precio unitario: "))
        nuevo_total = nueva_cantidad * nuevo_precio

        ws.cell(row=fila_encontrada, column=2, value=nuevo_producto)
        ws.cell(row=fila_encontrada, column=3, value=nueva_cantidad)
        ws.cell(row=fila_encontrada, column=4, value=nuevo_precio)
        ws.cell(row=fila_encontrada, column=5, value=nuevo_total)

        wb.save(archivo)
        print("Cotización actualizada correctamente.")
    else:
        print("No se encontró la cotización para ese cliente.")
        
def eliminar_cotizacion():
    cliente = input("Ingrese el nombre del cliente a eliminar: ")
    cliente_normalizado = normalizar_texto(cliente)

    wb = load_workbook(archivo)
    ws = wb.active
    fila_a_eliminar = None

    for i, fila in enumerate(ws.iter_rows(min_row=2), start=2):
        nombre_archivo = normalizar_texto(str(fila[0].value))
        if nombre_archivo == cliente_normalizado:
            fila_a_eliminar = i
            break

    if fila_a_eliminar:
        ws.delete_rows(fila_a_eliminar)
        wb.save(archivo)
        print(f"Cotización de '{cliente}' eliminada correctamente.")
    else:
        print("No se encontró la cotización del cliente.")

def vista_previa():
    cliente = input("Ingrese el nombre del cliente: ")
    cliente_normalizado = normalizar_texto(cliente)

    wb = load_workbook(archivo)
    ws = wb.active

    encontrado = False
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre_archivo = normalizar_texto(fila[0])
        if nombre_archivo == cliente_normalizado:
            print("\n📄 Vista previa de la cotización")
            print("-" * 40)
            print(f"Cliente       : {fila[0]}")
            print(f"Producto      : {fila[1]}")
            print(f"Cantidad      : {fila[2]}")
            print(f"Precio unitario: {fila[3]}")
            print(f"Total         : {fila[4]}")
            print("-" * 40)
            encontrado = True
            break

    if not encontrado:
        print("No se encontró ninguna cotización para ese cliente.")
        
def exportar_cotizacion():
    cliente = input("Ingrese el nombre del cliente para exportar: ")
    cliente_normalizado = normalizar_texto(cliente)
    cliente_archivo = limpiar_nombre_archivo(cliente_normalizado)

    wb_original = load_workbook(archivo)
    ws_original = wb_original.active

    for fila in ws_original.iter_rows(min_row=2, values_only=True):
        if normalizar_texto(fila[0]) == cliente_normalizado:
            wb_nuevo = Workbook()
            ws_nuevo = wb_nuevo.active
            ws_nuevo.title = "Cotización"
            ws_nuevo.append(["Cliente", "Producto", "Cantidad", "Precio Unitario", "Total"])
            ws_nuevo.append(fila)

            nombre_archivo = f"/storage/emulated/0/DojoHizoka/generador_cotizaciones/cotizacion_{cliente_archivo}.xlsx"
            wb_nuevo.save(nombre_archivo)
            print(f"Cotización exportada como:\n{nombre_archivo}")
            return

    print("No se encontró ninguna cotización para ese cliente.")